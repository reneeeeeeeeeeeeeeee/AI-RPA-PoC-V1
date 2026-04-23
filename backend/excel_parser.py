"""
Excel Parser
Liest Excel-Dateien, normalisiert Spalten, schreibt Ergebnisse zurück
"""

from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


# Bekannte Spalten-Aliases → Normname
COLUMN_ALIASES = {
    "lieferant": "Lieferant",
    "supplier":  "Lieferant",
    "vendor":    "Lieferant",
    "kreditor":  "Lieferant",

    "artikel":   "Artikel",
    "material":  "Artikel",
    "item":      "Artikel",
    "produkt":   "Artikel",
    "ref":       "Artikel",

    "menge":     "Menge",
    "quantity":  "Menge",
    "qty":       "Menge",
    "anzahl":    "Menge",

    "kostenstelle":    "Kostenstelle",
    "cost center":     "Kostenstelle",
    "kostenstellen":   "Kostenstelle",
    "kst":             "Kostenstelle",

    "prioritaet":  "Prioritaet",
    "priorität":   "Prioritaet",
    "priority":    "Prioritaet",
    "dringlichkeit": "Prioritaet",

    "preis":   "Preis",
    "price":   "Preis",
    "betrag":  "Preis",
}


class ExcelParser:

    def analyze(self, file_path: str) -> dict:
        df = pd.read_excel(file_path, dtype=str)
        df = df.fillna("")

        # Spalten normalisieren
        column_map = {}
        for col in df.columns:
            normalized = COLUMN_ALIASES.get(col.strip().lower())
            if normalized:
                column_map[col] = normalized

        # DataFrame umbenennen
        df = df.rename(columns=column_map)

        rows = []
        for i, row in df.iterrows():
            r = row.to_dict()
            r["_row"] = i + 2  # Excel-Zeile (1-basiert + Header)
            rows.append(r)

        return {
            "file_path": file_path,
            "total": len(rows),
            "rows": rows,
            "column_map": column_map,
            "columns": list(df.columns),
        }

    def write_results(self, file_path: str, results: list, errors: list):
        if not file_path or not Path(file_path).exists():
            return

        wb = load_workbook(file_path)
        ws = wb.active

        # Header-Zeile prüfen / Spalten hinzufügen
        headers = [cell.value for cell in ws[1]]
        result_cols = ["Status", "Beleg-Nr.", "Zeitstempel", "Fehler"]
        for col_name in result_cols:
            if col_name not in headers:
                ws.cell(row=1, column=len(headers) + 1, value=col_name)
                headers.append(col_name)

        col_idx = {h: i + 1 for i, h in enumerate(headers)}

        # Ergebnisse schreiben
        for r in results:
            row = r.get("excel_row")
            if not row:
                continue
            ws.cell(row=row, column=col_idx["Status"],     value="OK")
            ws.cell(row=row, column=col_idx["Beleg-Nr."],  value=r.get("beleg_nr", ""))
            ws.cell(row=row, column=col_idx["Zeitstempel"],value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        for e in errors:
            row = e.get("row")
            if not row:
                continue
            ws.cell(row=row, column=col_idx["Status"], value="FEHLER")
            ws.cell(row=row, column=col_idx["Fehler"],  value=e.get("reason", ""))
            ws.cell(row=row, column=col_idx["Zeitstempel"], value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        out_path = str(file_path).replace(".xlsx", "_ergebnis.xlsx")
        wb.save(out_path)
        print(f"[ExcelParser] Ergebnis gespeichert: {out_path}")
        return out_path
